# exe 빌드: pyinstaller --onefile --windowed --name "DRMforking" drm_gui.py
# python -m pip install pywin32
# 한글 및 MS오피스 설치된 컴퓨터에서 사용 가능

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import win32com.client as win32
import os
import shutil
import tempfile
import threading
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# 지원 확장자 → 처리 타입 매핑
EXTENSION_MAP = {
    ".hwp": "hwp",
    ".hwpx": "hwp",
    ".xls": "excel",
    ".xlsx": "excel",
    ".doc": "word",
    ".docx": "word",
    ".ppt": "ppt",
    ".pptx": "ppt",
    ".txt": "txt",
    ".pdf": "pdf",
    ".jpg": "image",
    ".jpeg": "image",
    ".png": "image",
}

FILETYPES = [
    ("지원 파일", "*.hwp;*.hwpx;*.xls;*.xlsx;*.doc;*.docx;*.ppt;*.pptx;*.txt;*.pdf;*.jpg;*.jpeg;*.png"),
    ("한글 파일", "*.hwp;*.hwpx"),
    ("엑셀 파일", "*.xls;*.xlsx"),
    ("워드 파일", "*.doc;*.docx"),
    ("파워포인트 파일", "*.ppt;*.pptx"),
    ("텍스트 파일", "*.txt"),
    ("PDF 파일", "*.pdf"),
    ("이미지 파일", "*.jpg;*.jpeg;*.png"),
    ("모든 파일", "*.*"),
]


class DrmApp:
    def __init__(self, root):
        self.root = root
        self.root.title("DRM 문서 도구")
        self.root.geometry("700x500")
        self.root.resizable(True, True)

        self._build_ui()

    def _build_ui(self):
        # 파일 선택 영역
        frame_top = tk.Frame(self.root, padx=10, pady=10)
        frame_top.pack(fill=tk.X)

        tk.Label(frame_top, text="파일 경로:").pack(side=tk.LEFT)

        self.entry_path = tk.Entry(frame_top)
        self.entry_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 5))

        btn_browse = tk.Button(frame_top, text="찾아보기", command=self._browse_file)
        btn_browse.pack(side=tk.LEFT)

        # 버튼 영역
        frame_btn = tk.Frame(self.root, padx=10, pady=5)
        frame_btn.pack(fill=tk.X)

        self.btn_open = tk.Button(
            frame_btn, text="열기", width=15, command=self._on_open
        )
        self.btn_open.pack(side=tk.LEFT, padx=(0, 10))

        self.btn_save = tk.Button(
            frame_btn, text="새로저장 (.iso)", width=15, command=self._on_save_iso
        )
        self.btn_save.pack(side=tk.LEFT)

        # 로그 영역
        frame_log = tk.Frame(self.root, padx=10, pady=5)
        frame_log.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame_log, text="로그:").pack(anchor=tk.W)
        self.log_text = scrolledtext.ScrolledText(frame_log, height=15, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def _browse_file(self):
        path = filedialog.askopenfilename(filetypes=FILETYPES)
        if path:
            self.entry_path.delete(0, tk.END)
            self.entry_path.insert(0, path)

    def _log(self, msg):
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def _set_buttons(self, enabled):
        state = tk.NORMAL if enabled else tk.DISABLED
        self.btn_open.configure(state=state)
        self.btn_save.configure(state=state)

    def _get_file_info(self):
        path = self.entry_path.get().strip()
        if not path:
            messagebox.showwarning("경고", "파일을 선택하세요.")
            return None, None
        abs_path = os.path.abspath(path)
        if not os.path.isfile(abs_path):
            messagebox.showerror("오류", f"파일이 존재하지 않습니다:\n{abs_path}")
            return None, None
        ext = os.path.splitext(abs_path)[1].lower()
        doc_type = EXTENSION_MAP.get(ext)
        if doc_type is None:
            messagebox.showerror("오류", f"지원하지 않는 확장자입니다: {ext}")
            return None, None
        return abs_path, doc_type

    # ── 열기 전용 ──────────────────────────────────────────

    def _on_open(self):
        abs_path, doc_type = self._get_file_info()
        if abs_path is None:
            return
        self._set_buttons(False)
        threading.Thread(target=self._run_open, args=(abs_path, doc_type), daemon=True).start()

    def _run_open(self, abs_path, doc_type):
        try:
            if doc_type == "hwp":
                self._open_hwp(abs_path)
            elif doc_type == "excel":
                self._open_excel(abs_path)
            elif doc_type == "word":
                self._open_word(abs_path)
            elif doc_type == "ppt":
                self._open_ppt(abs_path)
            elif doc_type == "txt":
                self._open_txt(abs_path)
            elif doc_type == "pdf":
                self._open_pdf(abs_path)
            elif doc_type == "image":
                self._open_image(abs_path)
        except Exception as e:
            self._log(f"[에러] {e}")
        finally:
            self.root.after(0, self._set_buttons, True)

    def _open_hwp(self, abs_path):
        self._log(f"[HWP] 한글 실행 중...")
        hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
        hwp.RegisterModule("FilePathCheckerModule", "FilePathChecker")
        hwp.XHwpWindows.Item(0).Visible = True
        if hwp.Open(abs_path):
            self._log(f"[HWP] 파일 열기 성공: {abs_path}")
        else:
            self._log(f"[HWP] 파일 열기 실패: {abs_path}")
        # 열기 모드에서는 종료하지 않음 (사용자가 직접 확인)

    def _open_excel(self, abs_path):
        self._log(f"[Excel] 엑셀 실행 중...")
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = True
        excel.DisplayAlerts = True
        excel.Workbooks.Open(abs_path)
        self._log(f"[Excel] 파일 열기 성공: {abs_path}")

    def _open_word(self, abs_path):
        self._log(f"[Word] 워드 실행 중...")
        word = win32.gencache.EnsureDispatch("Word.Application")
        word.Visible = True
        word.Documents.Open(abs_path)
        self._log(f"[Word] 파일 열기 성공: {abs_path}")

    def _open_ppt(self, abs_path):
        self._log(f"[PPT] 파워포인트 실행 중...")
        ppt_app = win32.gencache.EnsureDispatch("PowerPoint.Application")
        ppt_app.Visible = True
        ppt_app.Presentations.Open(abs_path, WithWindow=True)
        self._log(f"[PPT] 파일 열기 성공: {abs_path}")

    def _open_txt(self, abs_path):
        self._log(f"[TXT] 워드로 실행 중...")
        word = win32.gencache.EnsureDispatch("Word.Application")
        word.Visible = True
        word.Documents.Open(abs_path)
        self._log(f"[TXT] 파일 열기 성공: {abs_path}")

    def _open_pdf(self, abs_path):
        self._log(f"[PDF] 기본 뷰어로 실행 중...")
        os.startfile(abs_path)
        self._log(f"[PDF] 파일 열기 성공: {abs_path}")

    def _open_image(self, abs_path):
        self._log(f"[IMG] 기본 뷰어로 실행 중...")
        os.startfile(abs_path)
        self._log(f"[IMG] 파일 열기 성공: {abs_path}")

    # ── .iso 저장 ──────────────────────────────────────────

    def _on_save_iso(self):
        abs_path, doc_type = self._get_file_info()
        if abs_path is None:
            return
        self._set_buttons(False)
        threading.Thread(target=self._run_save_iso, args=(abs_path, doc_type), daemon=True).start()

    def _run_save_iso(self, abs_path, doc_type):
        try:
            save_path = os.path.splitext(abs_path)[0] + ".iso"
            if doc_type == "hwp":
                self._save_hwp_iso(abs_path, save_path)
            elif doc_type == "excel":
                self._save_excel_iso(abs_path, save_path)
            elif doc_type == "word":
                self._save_word_iso(abs_path, save_path)
            elif doc_type == "ppt":
                self._save_ppt_iso(abs_path, save_path)
            elif doc_type == "txt":
                self._save_txt_iso(abs_path, save_path)
            elif doc_type == "pdf":
                self._save_pdf_iso(abs_path, save_path)
            elif doc_type == "image":
                self._save_image_iso(abs_path, save_path)
        except Exception as e:
            self._log(f"[에러] {e}")
        finally:
            self.root.after(0, self._set_buttons, True)

    def _save_hwp_iso(self, abs_path, save_path):
        self._log(f"[HWP] 한글 실행 중...")
        hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
        hwp.RegisterModule("FilePathCheckerModule", "FilePathChecker")
        hwp.XHwpWindows.Item(0).Visible = True
        if not hwp.Open(abs_path):
            self._log(f"[HWP] 파일 열기 실패: {abs_path}")
            hwp.Quit()
            return
        self._log(f"[HWP] 파일 열기 성공")
        if hwp.SaveAs(save_path, "HWP"):
            self._log(f"[HWP] ISO 저장 성공: {save_path}")
        else:
            self._log(f"[HWP] ISO 저장 실패")
        hwp.Quit()

    def _xl_color(self, color_int):
        """Excel BGR color → AARRGGBB hex (openpyxl 형식)"""
        try:
            c = int(color_int)
            if c <= 0:
                return None
            return f"FF{c & 0xFF:02X}{(c >> 8) & 0xFF:02X}{(c >> 16) & 0xFF:02X}"
        except (TypeError, ValueError):
            return None

    def _xl_border_side(self, border):
        """Excel Border → openpyxl Side"""
        try:
            if border.LineStyle is None or border.LineStyle == -4142:
                return Side()
            style_map = {1: 'hair', 2: 'thin', -4138: 'medium', 4: 'thick'}
            style = style_map.get(border.Weight, 'thin')
            color = self._xl_color(border.Color) or 'FF000000'
            return Side(style=style, color=color)
        except Exception:
            return Side()

    def _save_excel_iso(self, abs_path, save_path):
        excel = None
        try:
            self._log("[Excel] 엑셀 실행 중...")
            excel = win32.gencache.EnsureDispatch("Excel.Application")
            excel.Visible = True
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(abs_path)
            self._log("[Excel] 파일 열기 성공")

            wb_new = openpyxl.Workbook()
            wb_new.remove(wb_new.active)

            for sheet_idx in range(1, wb.Sheets.Count + 1):
                ws = wb.Sheets(sheet_idx)
                ws_new = wb_new.create_sheet(ws.Name)

                used = ws.UsedRange
                if used is None:
                    continue
                values = used.Value
                if values is None:
                    continue
                if not isinstance(values, tuple):
                    values = ((values,),)
                elif not isinstance(values[0], tuple):
                    values = (values,)

                sr = used.Row
                sc = used.Column
                merged_done = set()

                self._log(f"[Excel] '{ws.Name}' 시트 처리 중... ({len(values)}행)")

                h_map = {-4131: 'left', -4108: 'center', -4152: 'right', -4130: 'justify'}
                v_map = {-4160: 'top', -4108: 'center', -4107: 'bottom'}

                for r, row in enumerate(values):
                    for c, val in enumerate(row):
                        cr = sr + r
                        cc = sc + c
                        dst = ws_new.cell(row=cr, column=cc)
                        if val is not None:
                            dst.value = val

                        try:
                            src = ws.Cells(cr, cc)

                            # 숫자 서식 (한국어 "G/표준" → "General" 변환)
                            nf = str(src.NumberFormat) if src.NumberFormat else ''
                            nf = nf.replace('G/표준', 'General')
                            if nf:
                                dst.number_format = nf

                            # 폰트
                            sf = src.Font
                            dst.font = Font(
                                name=sf.Name or None,
                                size=sf.Size or None,
                                bold=bool(sf.Bold) if sf.Bold is not None else False,
                                italic=bool(sf.Italic) if sf.Italic is not None else False,
                                color=self._xl_color(sf.Color),
                            )

                            # 배경색
                            interior = src.Interior
                            if interior.Pattern and interior.Pattern != -4142:
                                fc = self._xl_color(interior.Color)
                                if fc:
                                    dst.fill = PatternFill('solid', fgColor=fc)

                            # 정렬
                            dst.alignment = Alignment(
                                horizontal=h_map.get(src.HorizontalAlignment),
                                vertical=v_map.get(src.VerticalAlignment),
                                wrap_text=bool(src.WrapText) if src.WrapText else False,
                            )

                            # 테두리
                            dst.border = Border(
                                left=self._xl_border_side(src.Borders(7)),
                                top=self._xl_border_side(src.Borders(8)),
                                bottom=self._xl_border_side(src.Borders(9)),
                                right=self._xl_border_side(src.Borders(10)),
                            )

                            # 병합 셀
                            if src.MergeCells:
                                ma = src.MergeArea.Address.replace('$', '')
                                if ma not in merged_done:
                                    merged_done.add(ma)
                                    ws_new.merge_cells(ma)
                        except Exception:
                            pass

                # 열 너비
                for c in range(used.Columns.Count):
                    try:
                        cc = sc + c
                        w = ws.Columns(cc).ColumnWidth
                        if w:
                            ws_new.column_dimensions[get_column_letter(cc)].width = float(w) + 1
                    except Exception:
                        pass

                # 행 높이
                for r in range(used.Rows.Count):
                    try:
                        cr = sr + r
                        h = ws.Rows(cr).RowHeight
                        if h:
                            ws_new.row_dimensions[cr].height = float(h)
                    except Exception:
                        pass

            wb.Close(False)
            self._log("[Excel] openpyxl로 저장 중...")
            wb_new.save(save_path)
            self._log(f"[Excel] ISO 저장 성공: {save_path}")
        finally:
            if excel:
                excel.Quit()

    def _save_word_iso(self, abs_path, save_path):
        word = None
        try:
            self._log(f"[Word] 워드 실행 중...")
            word = win32.gencache.EnsureDispatch("Word.Application")
            word.Visible = True
            doc = word.Documents.Open(abs_path)
            self._log(f"[Word] 파일 열기 성공")
            doc.SaveAs2(save_path, 16)  # 16 = wdFormatDocumentDefault (.docx 포맷)
            self._log(f"[Word] ISO 저장 성공: {save_path}")
            doc.Close(False)
        finally:
            if word:
                word.Quit()

    def _save_ppt_iso(self, abs_path, save_path):
        ppt_app = None
        try:
            self._log(f"[PPT] 파워포인트 실행 중...")
            ppt_app = win32.gencache.EnsureDispatch("PowerPoint.Application")
            ppt_app.Visible = True
            presentation = ppt_app.Presentations.Open(abs_path, WithWindow=True)
            self._log(f"[PPT] 파일 열기 성공")
            presentation.SaveAs(save_path, 24)  # 24 = ppSaveAsOpenXMLPresentation (.pptx 포맷)
            self._log(f"[PPT] ISO 저장 성공: {save_path}")
            presentation.Close()
        finally:
            if ppt_app:
                ppt_app.Quit()


    def _save_txt_iso(self, abs_path, save_path):
        word = None
        try:
            self._log(f"[TXT] 워드 실행 중...")
            word = win32.gencache.EnsureDispatch("Word.Application")
            word.Visible = True
            doc = word.Documents.Open(abs_path)
            self._log(f"[TXT] 파일 열기 성공")
            doc.SaveAs2(save_path, 2)  # 2 = wdFormatText (plain text)
            self._log(f"[TXT] ISO 저장 성공: {save_path}")
            doc.Close()
        finally:
            if word:
                word.Quit()

    def _save_pdf_iso(self, abs_path, save_path):
        word = None
        try:
            self._log(f"[PDF] 워드 실행 중...")
            word = win32.gencache.EnsureDispatch("Word.Application")
            word.Visible = True
            doc = word.Documents.Open(abs_path)
            self._log(f"[PDF] 파일 열기 성공")
            doc.ExportAsFixedFormat(save_path, 17)  # 17 = wdExportFormatPDF
            self._log(f"[PDF] ISO 저장 성공: {save_path}")
            doc.Close(False)
        finally:
            if word:
                word.Quit()

    def _save_image_iso(self, abs_path, save_path):
        ppt_app = None
        try:
            self._log(f"[IMG] 파워포인트 실행 중...")
            ppt_app = win32.gencache.EnsureDispatch("PowerPoint.Application")
            ppt_app.Visible = True
            pres = ppt_app.Presentations.Add()
            slide = pres.Slides.Add(1, 12)  # 12 = ppLayoutBlank
            pic = slide.Shapes.AddPicture(
                abs_path, False, True, 0, 0
            )
            # 슬라이드 크기를 이미지에 맞춤
            pres.PageSetup.SlideWidth = pic.Width
            pres.PageSetup.SlideHeight = pic.Height
            pic.Left = 0
            pic.Top = 0
            # 임시 파일로 Export 후 .iso로 이동
            ext = os.path.splitext(abs_path)[1].lower()
            fmt = "JPG" if ext in (".jpg", ".jpeg") else "PNG"
            tmp_ext = ".jpg" if fmt == "JPG" else ".png"
            with tempfile.NamedTemporaryFile(suffix=tmp_ext, delete=False) as tmp:
                tmp_path = tmp.name
            slide.Export(tmp_path, fmt)
            shutil.move(tmp_path, save_path)
            self._log(f"[IMG] ISO 저장 성공: {save_path}")
            pres.Saved = True
            pres.Close()
        finally:
            if ppt_app:
                ppt_app.Quit()


if __name__ == "__main__":
    root = tk.Tk()
    app = DrmApp(root)
    root.mainloop()
